from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter


def apply_thin_borders(
    worksheet,
    cell_range: str,
) -> None:
    """
    Apply thin borders to every cell
    in a worksheet range.
    """

    thin_side = Side(
        style="thin",
        color="D9E2F3",
    )

    thin_border = Border(
        left=thin_side,
        right=thin_side,
        top=thin_side,
        bottom=thin_side,
    )

    for row in worksheet[cell_range]:
        for cell in row:
            cell.border = thin_border


def auto_size_columns(
    worksheet,
    minimum_width: int = 12,
    maximum_width: int = 35,
) -> None:
    """
    Resize worksheet columns based on
    the longest value in each column.
    """

    for column_cells in worksheet.columns:
        column_letter = get_column_letter(
            column_cells[0].column
        )

        longest_value = 0

        for cell in column_cells:
            if cell.value is None:
                continue

            longest_value = max(
                longest_value,
                len(str(cell.value)),
            )

        adjusted_width = min(
            max(
                longest_value + 2,
                minimum_width,
            ),
            maximum_width,
        )

        worksheet.column_dimensions[
            column_letter
        ].width = adjusted_width

def build_executive_summary(
    worksheet,
    portfolio_summary: dict,
    title_fill,
    title_font,
    title_alignment,
    label_font,
    positive_font,
    negative_font,
) -> None:
    """Build and format the Executive Summary worksheet."""

    worksheet["A1"] = "Portfolio Analysis Report"
    worksheet.merge_cells("A1:B1")

    worksheet["A1"].fill = title_fill
    worksheet["A1"].font = title_font
    worksheet["A1"].alignment = title_alignment
    worksheet.row_dimensions[1].height = 24

    worksheet["A3"] = "Analysis Period"
    worksheet["A4"] = "Total Return"
    worksheet["A5"] = "Annualized Return"
    worksheet["A6"] = "Annualized Volatility"
    worksheet["A7"] = "Sharpe Ratio"
    worksheet["A8"] = "Maximum Drawdown"
    worksheet["A9"] = "Largest Contributor"
    worksheet["A10"] = "Largest Detractor"

    for row in range(3, 11):
        worksheet[f"A{row}"].font = label_font

    worksheet["B3"] = (
        f"{portfolio_summary['start_date']} to "
        f"{portfolio_summary['end_date']}"
    )
    worksheet["B4"] = portfolio_summary["total_return"]
    worksheet["B5"] = portfolio_summary["annualized_return"]
    worksheet["B6"] = portfolio_summary["annualized_volatility"]
    worksheet["B7"] = portfolio_summary["sharpe_ratio"]
    worksheet["B8"] = portfolio_summary["maximum_drawdown"]
    worksheet["B9"] = portfolio_summary["largest_contributor"]
    worksheet["B10"] = portfolio_summary["largest_detractor"]

    percentage_cells = ["B4", "B5", "B6", "B8"]
    performance_cells = ["B4", "B5", "B8"]

    for cell in percentage_cells:
        worksheet[cell].number_format = "0.00%"
        worksheet[cell].alignment = Alignment(
            horizontal="right",
        )

    for cell in performance_cells:
        if worksheet[cell].value < 0:
            worksheet[cell].font = negative_font
        else:
            worksheet[cell].font = positive_font

    worksheet["B7"].number_format = "0.00"
    worksheet["B7"].alignment = Alignment(
        horizontal="right",
    )

    apply_thin_borders(
        worksheet=worksheet,
        cell_range="A3:B10",
    )

    auto_size_columns(worksheet)
    worksheet.freeze_panes = "A3"
def build_holdings_sheet(
    worksheet,
    portfolio_weights: dict,
    total_contributions: dict,
    title_fill,
    title_font,
    title_alignment,
    header_fill,
    header_font,
    header_alignment,
    positive_font,
    negative_font,
) -> None:
    """Build and format the Holdings worksheet."""

    worksheet["A1"] = "Portfolio Holdings"

    worksheet["A3"] = "Ticker"
    worksheet["B3"] = "Weight"
    worksheet["C3"] = "Contribution"

    for row_number, (ticker, weight) in enumerate(
        portfolio_weights.items(),
        start=4,
    ):
        worksheet[f"A{row_number}"] = ticker
        worksheet[f"B{row_number}"] = weight
        worksheet[f"C{row_number}"] = (
            total_contributions[ticker]
        )

    worksheet.merge_cells("A1:C1")

    worksheet["A1"].fill = title_fill
    worksheet["A1"].font = title_font
    worksheet["A1"].alignment = title_alignment
    worksheet.row_dimensions[1].height = 24

    for cell in worksheet[3]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    last_holdings_row = 3 + len(portfolio_weights)

    for row_number in range(4, last_holdings_row + 1):
        worksheet[f"B{row_number}"].number_format = "0.00%"
        worksheet[f"C{row_number}"].number_format = "0.00%"

        contribution_value = worksheet[
            f"C{row_number}"
        ].value

        if contribution_value < 0:
            worksheet[f"C{row_number}"].font = negative_font
        else:
            worksheet[f"C{row_number}"].font = positive_font

    apply_thin_borders(
        worksheet=worksheet,
        cell_range=f"A3:C{last_holdings_row}",
    )

    auto_size_columns(worksheet)
    worksheet.freeze_panes = "A4"
def build_performance_sheet(
    worksheet,
    processed_data,
    title_fill,
    title_font,
    title_alignment,
    header_fill,
    header_font,
    header_alignment,
) -> None:
    """Build and format the Performance worksheet."""

    worksheet["A1"] = "Portfolio Performance"

    worksheet["A3"] = "Date"
    worksheet["B3"] = "Daily Return"
    worksheet["C3"] = "Cumulative Return"
    worksheet["D3"] = "Portfolio Value"
    worksheet["E3"] = "Drawdown"

    for row_number, (date, row_data) in enumerate(
        processed_data.sort_index().iterrows(),
        start=4,
    ):
        worksheet[f"A{row_number}"] = date.to_pydatetime()
        worksheet[f"B{row_number}"] = row_data[
            "Portfolio Return"
        ]
        worksheet[f"C{row_number}"] = row_data[
            "Portfolio Cumulative Return"
        ]
        worksheet[f"D{row_number}"] = row_data[
            "Portfolio Value"
        ]
        worksheet[f"E{row_number}"] = row_data[
            "Drawdown"
        ]

    worksheet.merge_cells("A1:E1")

    worksheet["A1"].fill = title_fill
    worksheet["A1"].font = title_font
    worksheet["A1"].alignment = title_alignment
    worksheet.row_dimensions[1].height = 24

    for cell in worksheet[3]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    last_performance_row = 3 + len(processed_data)

    for row_number in range(4, last_performance_row + 1):
        worksheet[f"A{row_number}"].number_format = "yyyy-mm-dd"
        worksheet[f"B{row_number}"].number_format = "0.00%"
        worksheet[f"C{row_number}"].number_format = "0.00%"
        worksheet[f"D{row_number}"].number_format = "0.00"
        worksheet[f"E{row_number}"].number_format = "0.00%"

    apply_thin_borders(
        worksheet=worksheet,
        cell_range=f"A3:E{last_performance_row}",
    )

    auto_size_columns(worksheet)
    worksheet.freeze_panes = "A4"
def build_risk_sheet(
    worksheet,
    portfolio_summary: dict,
    title_fill,
    title_font,
    title_alignment,
    label_font,
    header_fill,
    header_font,
    header_alignment,
    negative_font,
) -> None:
    """Build and format the Risk worksheet."""

    worksheet["A1"] = "Portfolio Risk Analysis"

    worksheet["A3"] = "Metric"
    worksheet["B3"] = "Value"

    worksheet["A4"] = "Annualized Volatility"
    worksheet["A5"] = "Sharpe Ratio"
    worksheet["A6"] = "Maximum Drawdown"
    worksheet["A7"] = "Risk-Free Rate"
    worksheet["A8"] = "95% Historical VaR"
    worksheet["A9"] = "95% Historical CVaR"
    worksheet["A10"] = "99% Historical VaR"
    worksheet["A11"] = "99% Historical CVaR"
    worksheet["B4"] = portfolio_summary[
        
        
        "annualized_volatility"
    ]
    worksheet["B5"] = portfolio_summary[
        "sharpe_ratio"
    ]
    worksheet["B6"] = portfolio_summary[
        "maximum_drawdown"
    ]
    worksheet["B7"] = portfolio_summary[
        "risk_free_rate"
    ]
    worksheet["B8"] = portfolio_summary[
        "value_at_risk_95"
    ]
    worksheet["B9"] = portfolio_summary[
        "conditional_value_at_risk_95"
    ]
    worksheet["B10"] = portfolio_summary[
        "value_at_risk_99"
    ]
    worksheet["B11"] = portfolio_summary[
        "conditional_value_at_risk_99"
    ]
    
    worksheet.merge_cells("A1:B1")

    worksheet["A1"].fill = title_fill
    worksheet["A1"].font = title_font
    worksheet["A1"].alignment = title_alignment
    worksheet.row_dimensions[1].height = 24

    for cell in worksheet[3]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    for row_number in range(4, 12):
        worksheet[f"A{row_number}"].font = label_font

    worksheet["B4"].number_format = "0.00%"
    worksheet["B5"].number_format = "0.00"
    worksheet["B6"].number_format = "0.00%"
    worksheet["B7"].number_format = "0.00%"
    worksheet["B8"].number_format = "0.00%"
    worksheet["B9"].number_format = "0.00%"
    worksheet["B10"].number_format = "0.00%"
    worksheet["B11"].number_format = "0.00%"

    if worksheet["B6"].value < 0:
        worksheet["B6"].font = negative_font

    apply_thin_borders(
        worksheet=worksheet,
        cell_range="A3:B11",
    )

    auto_size_columns(worksheet)
    worksheet.freeze_panes = "A4"
def build_charts_sheet(
    worksheet,
    chart_file_paths: list[Path],
) -> None:
    """Add portfolio charts to the Charts worksheet."""

    chart_positions = [
        "A1",
        "A25",
        "A49",
        "A73",
        "J1",
        "J25",
        "J49",
        "J73",
        "A97",
    ]

    for chart_path, position in zip(
        chart_file_paths,
        chart_positions,
    ):
        image = Image(chart_path)
        image.width = 600
        image.height = 340
        worksheet.add_image(
            image,
            position,
        )
def build_optimization_sheet(
    worksheet,
    current_portfolio_statistics: dict,
    maximum_sharpe_portfolio: dict,
    minimum_volatility_portfolio: dict,
    current_weights: dict,
    optimization_start_date: str,
    optimization_end_date: str,
    maximum_weight: float,
    monte_carlo_summary: dict,
    title_fill,
    title_font,
    title_alignment,
    header_fill,
    header_font,
    header_alignment,
    label_font,
) -> None:
    """Build and format the Portfolio Optimization worksheet."""

    worksheet["A1"] = "Portfolio Optimization Analysis"
    worksheet.merge_cells("A1:D1")

    worksheet["A1"].fill = title_fill
    worksheet["A1"].font = title_font
    worksheet["A1"].alignment = title_alignment
    worksheet.row_dimensions[1].height = 24

    worksheet["A3"] = "Metric"
    worksheet["B3"] = "Current Portfolio"
    worksheet["C3"] = "Maximum Sharpe"
    worksheet["D3"] = "Minimum Volatility"

    for cell in worksheet[3]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    worksheet["A4"] = "Expected Annual Return"
    worksheet["A5"] = "Expected Annual Volatility"
    worksheet["A6"] = "Expected Sharpe Ratio"

    for row_number in range(4, 7):
        worksheet[f"A{row_number}"].font = label_font

    worksheet["B4"] = current_portfolio_statistics["return"]
    worksheet["B5"] = current_portfolio_statistics["volatility"]
    worksheet["B6"] = current_portfolio_statistics["sharpe_ratio"]

    worksheet["C4"] = maximum_sharpe_portfolio["return"]
    worksheet["C5"] = maximum_sharpe_portfolio["volatility"]
    worksheet["C6"] = maximum_sharpe_portfolio["sharpe_ratio"]

    worksheet["D4"] = minimum_volatility_portfolio["return"]
    worksheet["D5"] = minimum_volatility_portfolio["volatility"]
    worksheet["D6"] = minimum_volatility_portfolio["sharpe_ratio"]

    for column in ("B", "C", "D"):
        worksheet[f"{column}4"].number_format = "0.00%"
        worksheet[f"{column}5"].number_format = "0.00%"
        worksheet[f"{column}6"].number_format = "0.00"

    worksheet["A9"] = "Ticker"
    worksheet["B9"] = "Current Weight"
    worksheet["C9"] = "Maximum Sharpe Weight"
    worksheet["D9"] = "Minimum Volatility Weight"

    for cell in worksheet[9]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    maximum_sharpe_weights = maximum_sharpe_portfolio[
        "weights"
    ]
    minimum_volatility_weights = minimum_volatility_portfolio[
        "weights"
    ]

    for row_number, ticker in enumerate(
        current_weights,
        start=10,
    ):
        worksheet[f"A{row_number}"] = ticker
        worksheet[f"B{row_number}"] = current_weights[ticker]
        worksheet[f"C{row_number}"] = maximum_sharpe_weights[ticker]
        worksheet[f"D{row_number}"] = minimum_volatility_weights[ticker]

        for column in ("B", "C", "D"):
            worksheet[
                f"{column}{row_number}"
            ].number_format = "0.00%"

    last_weight_row = 9 + len(current_weights)

    worksheet["F3"] = "Optimization Settings"
    worksheet["F3"].fill = header_fill
    worksheet["F3"].font = header_font
    worksheet["F3"].alignment = header_alignment

    worksheet["F4"] = "Estimation Period"
    worksheet["F5"] = "Maximum Asset Weight"
    worksheet["F4"].font = label_font
    worksheet["F5"].font = label_font

    worksheet["G4"] = (
        f"{optimization_start_date} to "
        f"{optimization_end_date}"
    )
    worksheet["G5"] = maximum_weight
    worksheet["G5"].number_format = "0.00%"
    worksheet["F8"] = "Monte Carlo Forecast"
    worksheet["F8"].fill = header_fill
    worksheet["F8"].font = header_font
    worksheet["F8"].alignment = header_alignment

    worksheet["F9"] = "Mean Ending Value"
    worksheet["F10"] = "Median Ending Value"
    worksheet["F11"] = "5th Percentile"
    worksheet["F12"] = "95th Percentile"
    worksheet["F13"] = "Probability of Loss"
    worksheet["F14"] = "Expected Gain"
    worksheet["F15"] = "5th-Percentile Downside"

    for row_number in range(9, 16):
        worksheet[f"F{row_number}"].font = label_font

    worksheet["G9"] = monte_carlo_summary[
        "mean_ending_value"
    ]
    worksheet["G10"] = monte_carlo_summary[
        "median_ending_value"
    ]
    worksheet["G11"] = monte_carlo_summary[
        "fifth_percentile"
    ]
    worksheet["G12"] = monte_carlo_summary[
        "ninety_fifth_percentile"
    ]
    worksheet["G13"] = monte_carlo_summary[
        "probability_of_loss"
    ]
    worksheet["G14"] = monte_carlo_summary[
        "expected_gain"
    ]
    worksheet["G15"] = monte_carlo_summary[
        "downside_value_at_risk"
    ]

    for row_number in (9, 10, 11, 12, 14, 15):
        worksheet[f"G{row_number}"].number_format = "0.00"

    worksheet["G13"].number_format = "0.00%"

    apply_thin_borders(
        worksheet=worksheet,
        cell_range="A3:D6",
    )

    apply_thin_borders(
        worksheet=worksheet,
        cell_range=f"A9:D{last_weight_row}",
    )

    apply_thin_borders(
        worksheet=worksheet,
        cell_range="F3:G5",
    )
    apply_thin_borders(
        worksheet=worksheet,
        cell_range="F8:G15",
    )

    auto_size_columns(worksheet)
    worksheet.freeze_panes = "A3"

def create_portfolio_workbook(
    portfolio_summary: dict,
    portfolio_weights: dict,
    total_contributions: dict,
    processed_data: pd.DataFrame,
    chart_file_paths: list[Path],
    current_portfolio_statistics: dict,
    maximum_sharpe_portfolio: dict,
    minimum_volatility_portfolio: dict,
    optimization_start_date: str,
    optimization_end_date: str,
    maximum_weight: float,
    monte_carlo_summary: dict,
    output_folder: Path,
) -> Path:
    """
    Create an Excel workbook for the portfolio report.

    Args:
        portfolio_summary:
            Dictionary containing portfolio summary statistics.
        portfolio_weights:
            Portfolio weights by ticker.
        total_contributions:
            Total return contribution by ticker.
            processed_data:
            DataFrame containing processed portfolio data.
        chart_file_paths:
            Paths to the chart image files included in the workbook.
        output_folder:
            Folder where the workbook will be saved.

    Returns:
        Path to the saved workbook.
    """

    output_folder.mkdir(
        parents=True,
        exist_ok=True,
    )

    workbook = Workbook()

    summary_sheet = workbook.active
    summary_sheet.title = "Executive Summary"

    holdings_sheet = workbook.create_sheet(
        title="Holdings",
    )

    performance_sheet = workbook.create_sheet(
        title="Performance",
    )

    risk_sheet = workbook.create_sheet(
        title="Risk",
    )

    optimization_sheet = workbook.create_sheet(
        title="Optimization",
    )

    charts_sheet = workbook.create_sheet(
        title="Charts",
    )

    title_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78",
    )

    title_font = Font(
        color="FFFFFF",
        bold=True,
        size=16,
    )

    title_alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    label_font = Font(
        bold=True,
    )

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="D9EAF7",
    )

    header_font = Font(
        bold=True,
        color="1F1F1F",
    )

    header_alignment = Alignment(
        horizontal="center",
        vertical="center",
    )
    positive_font = Font(
        color="008000",
    )

    negative_font = Font(
        color="C00000",
    )
    build_executive_summary(
        worksheet=summary_sheet,
        portfolio_summary=portfolio_summary,
        title_fill=title_fill,
        title_font=title_font,
        title_alignment=title_alignment,
        label_font=label_font,
        positive_font=positive_font,
        negative_font=negative_font,
    )

    build_holdings_sheet(
        worksheet=holdings_sheet,
        portfolio_weights=portfolio_weights,
        total_contributions=total_contributions,
        title_fill=title_fill,
        title_font=title_font,
        title_alignment=title_alignment,
        header_fill=header_fill,
        header_font=header_font,
        header_alignment=header_alignment,
        positive_font=positive_font,
        negative_font=negative_font,
    )
    build_performance_sheet(
        worksheet=performance_sheet,
        processed_data=processed_data,
        title_fill=title_fill,
        title_font=title_font,
        title_alignment=title_alignment,
        header_fill=header_fill,
        header_font=header_font,
        header_alignment=header_alignment,
    )
    build_risk_sheet(
        worksheet=risk_sheet,
        portfolio_summary=portfolio_summary,
        title_fill=title_fill,
        title_font=title_font,
        title_alignment=title_alignment,
        label_font=label_font,
        header_fill=header_fill,
        header_font=header_font,
        header_alignment=header_alignment,
        negative_font=negative_font,
    )
    build_optimization_sheet(
        worksheet=optimization_sheet,
        current_portfolio_statistics=(
            current_portfolio_statistics
        ),
        maximum_sharpe_portfolio=(
            maximum_sharpe_portfolio
        ),
        minimum_volatility_portfolio=(
            minimum_volatility_portfolio
        ),
        current_weights=portfolio_weights,
        optimization_start_date=(
            optimization_start_date
        ),
                optimization_end_date=(
            optimization_end_date
        ),
        maximum_weight=maximum_weight,
        monte_carlo_summary=monte_carlo_summary,
        title_fill=title_fill,
        title_font=title_font,
        title_alignment=title_alignment,
        header_fill=header_fill,
        header_font=header_font,
        header_alignment=header_alignment,
        label_font=label_font,
    )
    build_charts_sheet(
        worksheet=charts_sheet,
        chart_file_paths=chart_file_paths,
    )

    workbook.save(
        output_folder / "portfolio_report.xlsx"
    )

    workbook.close()

    return (
        output_folder
        / "portfolio_report.xlsx"
    )